#!/usr/bin/python3
# -*- coding: utf-8 -*-

'   self-made utils for operating .xlsx files based on openpyxl ---- by smdsbz   '

from openpyxl import Workbook, load_workbook
from flask import flash
from datetime import datetime
import os, sqlite3
from globalvar import *



######## utils ########

def _read_test():
    ''' return table header line '''
    wb = load_workbook(os.path.join(FOLDER, 'template.xlsx'))
    data = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    return tuple([ content.value for content in data['A2':'M2'][0] ])

def _move_cursor(ws, name='something you have to mess up with'):
    '''
    return the index asked
    IF no match THEN return index-number of the adjacent empty row
    '''
    nameCells = ws['A'][2:] # 1: test-use; running: 2
    names = tuple(filter(lambda s: s and s.strip(), [ content.value for content in nameCells ]))  # thx 2 MichealLiao
    if name in names:
        return names.index(name) + 3 # 2: test-use
    else:
        return len(names) + 3

def newFile(title="测试测试", depart="其它", *, date=str(datetime.now())):
    '''
      derive a new .xlsx from ./score-sheets/template.xlsx
      (openpyxl Compatibility: always use MS Excel to generate the template.xlsx)
    '''
    filename = title + '.xlsx'
    dst = os.path.join(FOLDER, filename)
    if filename in os.listdir(FOLDER):
        flash("该表格已经存在！", category='error')
        return 0
    try:
        wb = load_workbook(os.path.join(FOLDER, 'template.xlsx'))
    except IOError:
        flash("表格模板损坏！<br>请联系管理员！", category='error')
        return 0
    else:
        with sqlite3.connect(INVENTORY) as database:
            try:
                cursor = database.execute("insert into score (title, date, depart) values ('%s', '%s', '%s')" % (title, date, depart))
            except sqlite3.IntegrityError:
                flash("该表格名称已被占用！", category='error')
                return 0
            else:
                database.commit()
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        # fill header
        ws.title = title
        ws['B1'].value, ws['F1'].value, ws['J1'].value = title, depart, date
        # import names
        with sqlite3.connect(DATABASE) as database:
            cursor = database.execute("select name from test where depart = '%s'" % depart)
            names = cursor.fetchall()
            print('hi?')
        for i in range(len(names)):
            ws['A'+str(i+3)].value = names[i][0]
        wb.save(dst)
        # register at inventory.db
        flash("成功创建表格！<br>请一次性填写完表格！", category='success')
        return 1

def write(filename, raw):
    '''
      write/update ONE person at a time
      raw[0] should be person's name
    '''
    dst = os.path.join(FOLDER, filename)
    try:
        wb = load_workbook(dst)
    except IOError:
        flash("写入表格错误！", category='error')
        return 0
    else:
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        cur = str(_move_cursor(ws, raw['name']))
        ordered = [
            raw['dim-self'], raw['act-self'], raw['act-num'],
            raw['dly-self'], raw['dly-act'], raw['mntr-dim'],
            raw['mntr-act'], raw['attd'], raw['bonus'],
            raw['total']
        ]
        for o, i in zip(ws['B'+cur:'K'+cur][0], ordered):
            o.value = int(i)
        wb.save(dst)
        # Flask's flash won't show until refreshing, use js instead
        # flash("成功写入 %s 的分数！" % raw['name'], 'success')
        return 1

def read(filename):
    '''
      returns a dict of * in the file
        format:
        {'name':{'B':1, 'C': 2, ..., 'K':10}, 'next person':{...}, ...}
    '''
    # TODO: 同名问题
    dst = os.path.join(FOLDER, filename)
    try:
        wb = load_workbook(dst)
    except IOError:
        print('IOError during read({})'.format(dst))
        flash("表格读取错误！", category='error')
        return []
    else:
        print('load successully!')
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        end = _move_cursor(ws)
        everyone = dict()
        for row in tuple(map(str, range(3, end))):  # get ('3', '4', '5', ..., '{end-1}')
            someone = dict()  # single person container
            for col in 'BCDEFGHIJK':
                if ws[col+row].value is None:
                    someone[col] = ''
                else:
                    someone[col] = ws[col+row].value
            everyone[ws['A'+row].value] = someone  # join the party
        return everyone


def getPerson(filename, name):
    '''
      return a person at a time, in:
        {'name':'xxx', 'dim-self':x, 'act-self':x, ...}
    '''
    dst = os.path.join(FOLDER, filename)
    try:
        wb = load_workbook(dst)
    except IOError:
        print('IOError during read({})'.format(dst))
        flash("表格读取错误!", category='error')
        return dict()
    else:
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        cur = str(_move_cursor(ws, name))
        person = {
            'name': ws['A'+cur].value,
            'dim-self': ws['B'+cur].value,
            'act-self': ws['C'+cur].value,
            'act-num': ws['D'+cur].value,
            'dly-self': ws['E'+cur].value,
            'dly-act': ws['F'+cur].value,
            'mntr-dim': ws['G'+cur].value,
            'mntr-act': ws['H'+cur].value,
            'attd': ws['I'+cur].value,
            'bonus': ws['J'+cur].value
        }
        return person


def delFile(filename):
    ''' delete a file and cancel its registration '''
    dst = os.path.join(FOLDER, filename)
    # first: cancel registration
    with sqlite3.connect(INVENTORY) as database:
        cursor = database.execute("delete from score where title = '%s'" % filename.rstrip('.xlsx'))
        # even if there is no such record in db, this does not raise an error
        database.commit()
    # second: delete file
    # doesn't matter if this fails
    #     for the new xlsx always rewrite the old one
    try:
        os.remove(dst)
    except FileNotFoundError:
        flash("该表格实体不存在！<br>数据库中存在非法记录！", category='error')
        return 0
    else:
        flash("表格 %s 删除成功！" % filename.rstrip('.xlsx'), category='success')
        return 1
